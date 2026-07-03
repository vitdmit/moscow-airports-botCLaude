#!/usr/bin/env python3
"""Извлечение порейсовой истории из Excel-файлов коллег (SVO/VKO/DME).
Выход: history_from_new_files.csv, day_totals.csv, qc_extract.csv,
       date_fixes.csv, dedup_log.csv"""
from __future__ import annotations
import csv, datetime as dt, glob, re, sys
from collections import Counter
import openpyxl

FILE_SAVED = dt.date(2026, 7, 2)
WEEKDAYS = {"пн","вт","ср","чт","пт","сб","вс","понедельник","вторник","среда","четверг","пятница","суббота","воскресенье"}
SVO_TERMINALS = {
    "130-133, 145-146 Автобусные(С)": "C", "140-144 Шайба(С)": "C",
    "124-129, 134-139(С)": "C", "117-121 гейты - В": "B", "14-17 гейты - Д": "D",
}
SKIP_SHEETS = {"Гейты терминал С","Гейты- терминал В","Гейты - терминал Д","итог гейты ВВЛ","ВВЛ Рейсы","итог гейты МВЛ","Гейты - D","Гейты -С","Гейты - Е","победа ВВЛ"}
CYR2LAT = str.maketrans({"С":"C","Е":"E","В":"B","А":"A","с":"C","е":"E","в":"B","а":"A","Д":"D","д":"D"})
SUMMARY_PAT = re.compile(r"^(итого|итог|%|приоритетн|активн|модель|направлени|авиакомпани|дата)", re.I)
MONTHS3 = {"янв":1,"фев":2,"мар":3,"апр":4,"мая":5,"май":5,"июн":6,"июл":7,"авг":8,"сен":9,"окт":10,"ноя":11,"дек":12}
STR_DATE_RE = re.compile(r"^(\d{1,2})[.\s]+([а-яА-Я]{3,})\.?(?:\s+(\d{2}|\d{4}))?$")
NUM_DATE_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$")

def clean_gate(header):
    s = header.split("(")[0].strip()
    if "-" in s and any(ch.isdigit() for ch in s):
        parts = [p for p in re.split(r"[,\s]+", s) if p]
        if len(parts) > 1 and all("-" in p or p.replace("А","").replace("A","").isdigit() for p in parts):
            s = ",".join(parts)
    m = re.match(r"^([A-Za-zА-Яа-я]?\d+)\s*,\s*\1[АA]$", s)
    if m: s = m.group(1)
    return s.translate(CYR2LAT).strip().rstrip(",")

def fmt_time(v):
    if v is None or v == "": return ""
    if isinstance(v, dt.datetime): return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, dt.time): return f"{v.hour:02d}:{v.minute:02d}"
    if isinstance(v, (int, float)):
        mins = round(float(v) % 1 * 24 * 60)
        return f"{mins // 60 % 24:02d}:{mins % 60:02d}"
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})[:.](\d{2})", s)
    return f"{int(m.group(1)) % 24:02d}:{m.group(2)}" if m else s[:5]

def classify_a(v):
    """Значение колонки А -> ('expl', date) | ('dm', (day,mon)) | None."""
    if isinstance(v, dt.datetime): v = v.date()
    if isinstance(v, dt.date):
        return ("dm", (v.day, v.month)) if v.year < 1990 else ("expl", v)
    if not isinstance(v, str): return None
    s = re.sub(r"\s+", " ", v.strip()).lower().rstrip(".")
    s = re.sub(r"\s+(пн|вт|ср|чт|пт|сб|вс)\.?$", "", s)
    if not s or s in WEEKDAYS or s.startswith("итог"): return None
    m = STR_DATE_RE.match(s)
    if m and m.group(2)[:3] in MONTHS3:
        day, mon = int(m.group(1)), MONTHS3[m.group(2)[:3]]
        if m.group(3):
            y = int(m.group(3)); y += 2000 if y < 100 else 0
            if y >= 1990:
                try: return ("expl", dt.date(y, mon, day))
                except ValueError: return None
        return ("dm", (day, mon))
    m = NUM_DATE_RE.match(s)
    if m:
        day, mon = int(m.group(1)), int(m.group(2))
        if not (1 <= mon <= 12 and 1 <= day <= 31): return None
        if m.group(3):
            y = int(m.group(3)); y += 2000 if y < 100 else 0
            if y >= 1990:
                try: return ("expl", dt.date(y, mon, day))
                except ValueError: return None
        return ("dm", (day, mon))
    return None

def _fit(day, mon, lo, hi):
    """Дата (day,mon) с годом так, чтобы lo <= d <= hi."""
    if lo is None: return None
    for y in (lo.year, lo.year + 1, lo.year - 1):
        try: d = dt.date(y, mon, day)
        except ValueError: continue
        if lo <= d <= hi: return d
    return None

def resolve_dates(ws, head_i, airport, sheet, fixes):
    raw = [(i, row[0] if row else None) for i, row in
           enumerate(ws.iter_rows(min_row=head_i+1, max_col=1, values_only=True), head_i+1)]
    items = []
    for row_i, v in raw:
        c = classify_a(v)
        if c: items.append((row_i, c[0], c[1], v))
    expl_seq = [(k, it) for k, it in enumerate(items) if it[1] == "expl"]

    resolved = {}
    prev = None
    pending = []   # ведущие 'dm' до первого якоря
    for k, (row_i, kind, val, rawv) in enumerate(items):
        where = f"{airport}/{sheet} стр.{row_i}"
        # следующая явная дата (сырая)
        nxt = None
        for kk, it in expl_seq:
            if kk > k: nxt = it[2]; break
        if kind == "dm":
            if prev is None:
                pending.append((row_i, val[0], val[1], rawv)); continue
            d = _fit(val[0], val[1], prev, prev + dt.timedelta(days=40))
            if d:
                resolved[row_i] = d; prev = d
                fixes.append((where, str(rawv).strip()[:20], d.isoformat(), "дата без года"))
            else:
                # месяц бит, день верен: ищем ближайший день с таким номером
                d2 = None
                for delta in range(1, 41):
                    c = prev + dt.timedelta(days=delta)
                    if c.day == val[0]:
                        d2 = c; break
                if d2:
                    resolved[row_i] = d2; prev = d2
                    fixes.append((where, str(rawv).strip()[:20], d2.isoformat(), "восстановлена по дню месяца"))
                else:
                    fixes.append((where, str(rawv).strip()[:20], "", "дата без года — не разрешена, строки уйдут к предыдущему дню"))
            continue
        d = val
        if prev is None:
            resolved[row_i] = d; prev = d; continue
        delta = (d - prev).days
        outlier = False
        if -5 <= delta <= 40:
            if delta > 7 and nxt is not None and nxt < d and -5 <= (nxt - prev).days <= 40:
                outlier = True    # скачок вперёд, следующая дата возвращается
        elif nxt is not None and abs((nxt - d).days) <= 7 and not (-5 <= delta <= 40):
            # целый блок со сдвигом (дубль-вставка или возврат) — принимаем как есть
            fixes.append((where, d.isoformat(), d.isoformat(), "непоследовательный блок — принят как есть"))
        else:
            outlier = True
        if outlier:
            hi = min(prev + dt.timedelta(days=40), nxt) if (nxt and nxt >= prev) else prev + dt.timedelta(days=40)
            cands = []
            for y in (prev.year, prev.year + 1):
                try: cands.append(dt.date(y, d.month, d.day))
                except ValueError: pass
            for mth in (prev.month, prev.month % 12 + 1):
                try: cands.append(dt.date(prev.year, mth, d.day))
                except ValueError: pass
            fixed = next((c for c in cands if prev <= c <= hi), None)
            if fixed:
                fixes.append((where, d.isoformat(), fixed.isoformat(), "опечатка года/месяца"))
                d = fixed
            else:
                fixes.append((where, d.isoformat(), d.isoformat(), "аномальная дата — оставил как есть"))
                resolved[row_i] = d
                continue   # prev не двигаем
        resolved[row_i] = d
        prev = d
    # ведущие 'dm' — обратный проход от первого якоря
    if pending:
        anchor = None
        for row_i in sorted(resolved):
            anchor = resolved[row_i]; break
        for p_row, p_day, p_mon, p_raw in reversed(pending):
            if anchor is None: break
            d = _fit(p_day, p_mon, anchor - dt.timedelta(days=40), anchor)
            if d is None:
                fixes.append((f"{airport}/{sheet} стр.{p_row}", str(p_raw)[:20], "", "дата без года в начале листа — не разрешена"))
                continue
            resolved[p_row] = d
            fixes.append((f"{airport}/{sheet} стр.{p_row}", str(p_raw)[:20], d.isoformat(), "дата без года (обратный проход)"))
            anchor = d
    return resolved

def parse_sheet(ws, airport, sheet, rows_out, totals_out, qc_out, fixes):
    head, head_i = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), 1):
        vals = ["" if c is None else str(c).strip() for c in row]
        if vals and vals[0].lower() == "дата":
            head, head_i = vals, i; break
    if head is None:
        print(f"  !! {airport}/{sheet}: нет строки заголовка — пропуск"); return
    gate_cols, total_cols = [], []
    seg_col, seg_name = None, ""
    for j, h in enumerate(head):
        if not h: continue
        hl = h.strip()
        if j == 0 or SUMMARY_PAT.match(hl):
            low = hl.lower()
            if low.startswith(("итого","итог")) and "ввл" not in low and "мвл" not in low and j > 0:
                total_cols.append((j, hl))
            if low.startswith("итого ввл"): seg_col, seg_name = j, "ВВЛ"
            if low.startswith("итого мвл"): seg_col, seg_name = j, "МВЛ"
            continue
        gate_cols.append((j, clean_gate(hl)))
    if not gate_cols:
        print(f"  !! {airport}/{sheet}: нет колонок гейтов"); return
    max_need = max(max(j for j,_ in gate_cols)+3, (seg_col or 0)+1, max((j for j,_ in total_cols), default=0)+1)

    resolved = resolve_dates(ws, head_i, airport, sheet, fixes)
    cur_date = None
    day_counts, day_totals = {}, {}
    orphans = []
    for row_i, row in enumerate(ws.iter_rows(min_row=head_i+1, max_col=max_need, values_only=True), head_i+1):
        got = resolved.get(row_i)
        if got is not None:
            cur_date = got
            if orphans:
                n_orph = 0
                for orow in orphans:
                    for r in orow:
                        r["flight_date"] = cur_date.isoformat()
                        rows_out.append(r); n_orph += 1
                day_counts[cur_date] = day_counts.get(cur_date, 0) + n_orph
                fixes.append((f"{airport}/{sheet}", "строки до первой даты", cur_date.isoformat(), f"{n_orph} рейсов приписано первому дню"))
                orphans = []
            tot, has = 0, False
            for j,_ in total_cols:
                v = row[j] if j < len(row) else None
                if isinstance(v, (int, float)): tot += int(v); has = True
            seg_v = None
            if seg_col is not None and seg_col < len(row):
                v = row[seg_col]
                if isinstance(v, (int, float)): seg_v = int(v)
            if has or seg_v is not None:
                # при дубль-блоках оставляем итог первого вхождения
                day_totals.setdefault(cur_date, (tot if has else None, seg_v))
        flights_here = []
        for j, gate in gate_cols:
            city = row[j] if j < len(row) else None
            t = row[j+1] if j+1 < len(row) else None
            al = row[j+2] if j+2 < len(row) else None
            city_s = "" if city is None else str(city).strip()
            if city_s.lower().startswith("итого"): continue
            t_s = fmt_time(t)
            if not city_s and not t_s: continue
            if not t_s and re.fullmatch(r"[\d\s.,%]+", city_s): continue
            al_s = "" if al is None else str(al).strip()
            if al_s.lower().startswith("итого") or re.fullmatch(r"[\d\s.,%]+", al_s): al_s = ""
            terminal = (SVO_TERMINALS.get(sheet, "") if airport == "SVO" else "A" if airport == "VKO" else (gate[0] if gate[:1].isalpha() else ""))
            flights_here.append({"airport": airport, "flight_date": "", "terminal": terminal, "gate": gate, "destination": city_s, "scheduled_time": t_s, "airline": al_s, "source_sheet": sheet, "is_future": False})
        if not flights_here: continue
        if cur_date is None:
            orphans.append(flights_here); continue
        for r in flights_here:
            r["flight_date"] = cur_date.isoformat()
            r["is_future"] = cur_date > FILE_SAVED
            rows_out.append(r)
        day_counts[cur_date] = day_counts.get(cur_date, 0) + len(flights_here)
    for d, n in sorted(day_counts.items()):
        their, _ = day_totals.get(d, (None, None))
        qc_out.append({"airport": airport, "source_sheet": sheet, "flight_date": d.isoformat(), "my_count": n, "their_total": their if their is not None else "", "diff": (n - their) if their is not None else ""})
    for d, (their, seg_v) in sorted(day_totals.items()):
        totals_out.append({"airport": airport, "source_sheet": sheet, "flight_date": d.isoformat(), "group_total": their if their is not None else "", "segment": seg_name, "segment_total": seg_v if seg_v is not None else ""})

def parse_pobeda(ws, totals_out):
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        d, n = row[0], row[1]
        if isinstance(d, (dt.datetime, dt.date)) and isinstance(n, (int, float)):
            dd = d.date() if isinstance(d, dt.datetime) else d
            totals_out.append({"airport": "VKO", "source_sheet": "победа ВВЛ", "flight_date": dd.isoformat(), "group_total": int(n), "segment": "ВВЛ", "segment_total": ""})

def dedupe_block_copies(rows, qc, dedup_log):
    """Схлопнуть точные дубль-блоки: если ВСЕ рейсы (лист, дата) продублированы
    одинаковое число раз (>=2) — оставить один комплект."""
    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        groups[(r["source_sheet"], r["flight_date"])].append(r)
    out = []
    deduped_days = set()
    for key, grp in groups.items():
        cnt = Counter((r["gate"], r["scheduled_time"], r["destination"], r["airline"]) for r in grp)
        mults = set(cnt.values())
        if len(mults) == 1 and mults != {1} and len(cnt) >= 3:
            m = mults.pop()
            seen = Counter()
            kept = []
            for r in grp:
                k = (r["gate"], r["scheduled_time"], r["destination"], r["airline"])
                seen[k] += 1
                if seen[k] <= cnt[k] // m:
                    kept.append(r)
            out.extend(kept)
            dedup_log.append({"source_sheet": key[0], "flight_date": key[1], "было": len(grp), "стало": len(kept), "кратность": m})
            deduped_days.add(key)
        else:
            out.extend(grp)
    if deduped_days:
        for q in qc:
            k = (q["source_sheet"], q["flight_date"])
            if k in deduped_days:
                q["my_count"] = q["my_count"] // 2 if isinstance(q["my_count"], int) else q["my_count"]
    return out

def main(src_dir=".", out_dir="."):
    files = {k: glob.glob(f"{src_dir}/*{k}*.xlsx")[0] for k in ["SVO","VKO","DME"]}
    rows, totals, qc, fixes, dedup_log = [], [], [], [], []
    for airport, path in files.items():
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if sn == "победа ВВЛ":
                parse_pobeda(wb[sn], totals); continue
            if sn in SKIP_SHEETS: continue
            parse_sheet(wb[sn], airport, sn, rows, totals, qc, fixes)
        wb.close()
        print(f"{airport}: накоплено {len(rows)}")
    rows = dedupe_block_copies(rows, qc, dedup_log)
    def dump(name, fieldnames, data, is_dict=True):
        with open(f"{out_dir}/{name}", "w", newline="", encoding="utf-8-sig") as f:
            if is_dict:
                w = csv.DictWriter(f, fieldnames=fieldnames); w.writeheader(); w.writerows(data)
            else:
                w = csv.writer(f); w.writerow(fieldnames); w.writerows(data)
    dump("history_from_new_files.csv", ["airport","flight_date","terminal","gate","destination","scheduled_time","airline","source_sheet","is_future"], rows)
    dump("day_totals.csv", ["airport","source_sheet","flight_date","group_total","segment","segment_total"], totals)
    dump("qc_extract.csv", ["airport","source_sheet","flight_date","my_count","their_total","diff"], qc)
    dump("date_fixes.csv", ["где","было","стало","причина"], fixes, is_dict=False)
    dump("dedup_log.csv", ["source_sheet","flight_date","было","стало","кратность"], dedup_log)
    print(f"Итого рейсов: {len(rows)}; исправлений дат: {len(fixes)}; дубль-дней схлопнуто: {len(dedup_log)}")

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else ".", sys.argv[2] if len(sys.argv) > 2 else ".")
