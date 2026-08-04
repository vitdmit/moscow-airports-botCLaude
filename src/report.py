"""Отчёт загрузки гейтов: ДВА базиса сравнения, без мешанины источников.

Проблема старой версии: доля точки считалась как «гейты точки / весь терминал»
из единой склейки история(коллеги) + бот. У коллег в терминале учтены только
свои гейты, у бота — все. При переходе июнь(коллеги)->июль(бот) знаменатель
скачком менялся, и доли рушились. Плюс склейка не дедуплицировалась (пересечение
конца июня считалось дважды).

Новая версия считает КАЖДУЮ метрику на одном согласованном базисе:

  1. «Весь аэропорт» — только данные бота (он собирает все гейты). Доступно с
     конца мая 2026. Доля точки = гейты точки / все рейсы терминала.

  2. «Наши гейты» — только гейты точек (набор TRACKED), данные бота. За
     месяцы до запуска бота (для год-к-году) берётся история коллег.
     Пересечение дат дедуплицируется в пользу бота. Доля точки = гейты точки
     / наши гейты терминала.

Так месяц-к-месяцу считается бот/бот (чисто), а год-к-году один раз проходит
через стык бот↔коллеги (прошлый год есть только у коллег).

Запуск:
    python -m src.report --to 2026-07
"""
from __future__ import annotations

import argparse
import re
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- точки Винегрет (аэропорт, терминал, спецификация гейтов) --------------
POINTS = [
    {"point": "АБ460",      "airport": "VKO", "terminal": "A", "gates": "11-12"},
    {"point": "Кофеин ВВЛ", "airport": "VKO", "terminal": "A", "gates": "13"},
    {"point": "Бад",        "airport": "VKO", "terminal": "A", "gates": "15-16"},
    {"point": "АБ99",       "airport": "VKO", "terminal": "A", "gates": "21-22"},
    {"point": "Ачарули",    "airport": "VKO", "terminal": "A", "gates": "23-24"},
    {"point": "Бир3/Кофеин МВЛ", "airport": "VKO", "terminal": "A", "gates": "24-25"},
    {"point": "Баттерфляй", "airport": "DME", "terminal": "C", "gates": "C5-C7"},
    {"point": "АБ131",      "airport": "DME", "terminal": "D", "gates": "D3-D8"},
    {"point": "Гурмэ Т2",   "airport": "DME", "terminal": "E", "gates": "E13,E14"},
    {"point": "Гурмэ В",    "airport": "SVO", "terminal": "B", "gates": "117-121"},
    {"point": "АБ600",      "airport": "SVO", "terminal": "D", "gates": "D14-D17"},
]

# ---- гейты, которые ведут коллеги (канонические ключи) ---------------------
# Знаменатель базиса «наши гейты». Меняются редко; при добавлении коллегами
# нового гейта — дописать сюда.
TRACKED = {
    "VKO": {"11", "12", "13", "15", "16", "18", "19", "20",
            "21", "22", "23", "24", "25"},
    "DME": {"C3", "C4", "C5", "C6", "C7", "C8", "C12", "C13", "C14", "C15",
            "C16", "C17", "C18", "C19", "D3", "D4", "D5", "D6", "D7", "D8",
            "D9", "D10", "D11", "D13", "D14", "D16", "D17",
            "E8", "E9", "E10", "E12", "E13", "E14"},
    "SVO": {"117", "118", "119", "120", "121",
            "124", "125", "126", "127", "128", "129", "130", "131", "132",
            "133", "134", "135", "136", "137", "138", "139", "140", "141",
            "142", "143", "144", "145", "146", "D14", "D15", "D16", "D17"},
}


def canon_gate(airport: str, terminal, gate) -> str | None:
    """Единый ключ гейта для бота и коллег.
    VKO: только цифры (11A и 11 -> '11', '08' -> '8').
    SVO: гейты терминала D 14-17 и 'D14' -> 'D14'; прочие цифры -> число;
         диапазонные ярлыки коллег ('124-129,134-139') остаются как есть.
    DME: 'C5','D3','E13' как есть (uppercase)."""
    if gate is None:
        return None
    g = str(gate).strip().upper()
    if g in ("", "—", "-", "NONE", "NAN"):
        return None
    t = str(terminal or "").strip().upper()
    if airport == "VKO":
        d = "".join(ch for ch in g if ch.isdigit())
        return str(int(d)) if d else None
    if airport == "SVO":
        m = re.match(r"^D0*(\d+)$", g)
        if m:
            return "D" + m.group(1)
        if g.isdigit():
            n = int(g)
            if t == "D" and n in (14, 15, 16, 17):
                return "D" + str(n)
            return str(n)
        return g
    return g


def _canon_num(prefix: str, num: int) -> str:
    return f"{prefix.upper()}{num}"


def expand_spec(spec: str) -> set[str]:
    """Спецификацию точки ('11-12','C5-C7','D14-D17','E13,E14','13') развернуть
    в множество канонических ключей."""
    out: set[str] = set()
    for tok in str(spec).split(","):
        tok = tok.strip().upper()
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            ma = re.match(r"^([A-ZА-Я]*)0*(\d+)[A-ZА-Я]?$", a)
            mb = re.match(r"^([A-ZА-Я]*)0*(\d+)[A-ZА-Я]?$", b)
            if not ma or not mb:
                continue
            prefix = ma.group(1) or mb.group(1)
            lo, hi = sorted([int(ma.group(2)), int(mb.group(2))])
            for n in range(lo, hi + 1):
                out.add(_canon_num(prefix, n))
        else:
            m = re.match(r"^([A-ZА-Я]*)0*(\d+)[A-ZА-Я]?$", tok)
            if m:
                out.add(_canon_num(m.group(1), int(m.group(2))))
    return out


POINT_GATES = {p["point"]: expand_spec(p["gates"]) for p in POINTS}

# ---- стили -----------------------------------------------------------------
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(FONT, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(FONT, bold=True, size=13, color="1F4E78")
BASE = Font(FONT, size=10)
BOLD = Font(FONT, bold=True, size=10)
GREY = Font(FONT, size=9, color="808080")
GREEN = Font(FONT, size=10, bold=True, color="006100")
RED = Font(FONT, size=10, bold=True, color="9C0006")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(THIN, THIN, THIN, THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def _prep(df: pd.DataFrame, dfrom, dto) -> pd.DataFrame:
    d = df.copy()
    d["flight_date"] = pd.to_datetime(d["flight_date"], errors="coerce")
    d = d[d["flight_date"].notna()].copy()
    d["ym"] = d["flight_date"].dt.strftime("%Y-%m")
    if dfrom:
        d = d[d["ym"] >= dfrom]
    if dto:
        d = d[d["ym"] <= dto]
    d["airport"] = d["airport"].astype(str).str.strip()
    d["terminal"] = d["terminal"].fillna("").astype(str).str.strip()
    if "src" not in d.columns:
        d["src"] = "daily"
    d["cgate"] = [canon_gate(a, t, g)
                  for a, t, g in zip(d["airport"], d["terminal"], d["gate"])]
    d["daykey"] = d["airport"] + "|" + d["flight_date"].dt.strftime("%Y-%m-%d")
    return d


def _whole(d: pd.DataFrame) -> pd.DataFrame:
    """Базис «весь аэропорт»: только бот."""
    return d[d["src"] == "daily"].copy()


def _tracked(d: pd.DataFrame) -> pd.DataFrame:
    """Базис «наши гейты»: приоритет у НАШИХ данных (бота), только гейты из
    TRACKED. Коллеги используются лишь для дат, которых у бота нет (месяцы до
    запуска бота — для год-к-году). Пересечение дат — в пользу бота (дедуп)."""
    bot_days = set(d.loc[d["src"] == "daily", "daykey"])
    keep = ~((d["src"] == "history") & (d["daykey"].isin(bot_days)))
    d = d[keep].copy()
    ok = [src == "history" or cg in TRACKED.get(ap, set())
          for ap, cg, src in zip(d["airport"], d["cgate"], d["src"])]
    return d[pd.Series(ok, index=d.index)].copy()


def _full_months(d: pd.DataFrame, min_days: int = 26) -> list[str]:
    g = d.groupby("ym")["flight_date"].apply(lambda s: s.dt.date.nunique())
    return sorted([m for m, n in g.items() if n >= min_days])


# ---- Сводка ----------------------------------------------------------------
def build_summary(wb, whole, tracked):
    ws = wb.active
    ws.title = "Сводка"
    ws.cell(1, 1, "Загрузка гейтов — сводка по месяцам, два базиса").font = TITLE_FONT
    ws.cell(2, 1, "«Весь аэропорт» — все рейсы (данные бота). «Наши гейты» — "
                  "только гейты точек, данные бота; за месяцы до запуска бота "
                  "берутся данные коллег (для год-к-году). Числа несопоставимы "
                  "между базисами, сопоставимы по месяцам внутри одного базиса.").font = GREY
    r = 4
    maxcols = 1
    for title, data in [("ВЕСЬ АЭРОПОРТ (бот, все гейты)", whole),
                        ("НАШИ ГЕЙТЫ (как у коллег)", tracked)]:
        months = sorted(set(data["ym"]))
        maxcols = max(maxcols, len(months))
        ws.cell(r, 1, title).font = BOLD
        r += 1
        _hdr(ws, r, ["Аэропорт"] + months)
        piv = (data.pivot_table(index="airport", columns="ym", values="cgate",
                                aggfunc="count", fill_value=0)
               .reindex(columns=months, fill_value=0))
        for ap in [a for a in ["SVO", "VKO", "DME"] if a in piv.index]:
            r += 1
            ws.cell(r, 1, ap).font = BOLD
            ws.cell(r, 1).border = BORDER
            for j, m in enumerate(months):
                c = ws.cell(r, 2 + j, int(piv.loc[ap, m]))
                c.font = BASE
                c.border = BORDER
                c.alignment = CENTER
        r += 2
    ws.column_dimensions["A"].width = 12
    for j in range(maxcols):
        ws.column_dimensions[get_column_letter(2 + j)].width = 9
    ws.freeze_panes = "B5"


# ---- Точки -----------------------------------------------------------------
def _series(data, point):
    """Для точки по месяцам: (кол-во рейсов на её гейтах, знаменатель терминала)."""
    ap, term = point["airport"], point["terminal"]
    gates = POINT_GATES[point["point"]]
    sub = data[(data["airport"] == ap) & (data["terminal"] == term)]
    num = sub[sub["cgate"].isin(gates)].groupby("ym").size()
    den = sub.groupby("ym").size()
    return num, den


def build_points(wb, whole, tracked):
    """Главный лист: загрузка НАШИХ гейтов у точек (базис коллег), динамика
    к прошлому месяцу и к тому же месяцу год назад."""
    ws = wb.create_sheet("Точки — динамика", 1)
    ws.cell(1, 1, "Точки Винегрет — загрузка наших гейтов, динамика").font = TITLE_FONT
    ws.cell(2, 1, "Данные бота, только гейты точек. Рейсы на гейтах точки за "
                  "последний полный месяц против прошлого месяца и того же месяца "
                  "год назад. Доля = гейты точки ÷ наши гейты терминала. Год-к-году "
                  "с прошлым годом (данные коллег): «н/д» — гейт тогда не вёлся, "
                  "переход бот↔коллеги может давать разовую ступеньку.").font = GREY

    full = _full_months(tracked)
    if not full:
        ws.cell(4, 1, "Недостаточно полных месяцев данных.").font = BASE
        return
    L = full[-1]
    P = full[-2] if len(full) >= 2 else None
    tym = set(tracked["ym"])
    YA = f"{int(L[:4]) - 1}-{L[5:]}"
    YA = YA if YA in tym else None

    r0 = 4
    _hdr(ws, r0, ["Точка", "Аэр", "Гейты точки",
                  (YA + "\nрейсов") if YA else "год назад",
                  (P + "\nрейсов") if P else "пр. месяц",
                  L + "\nрейсов", "Δ мес, %", "Δ год, %", "Доля " + L])
    ws.row_dimensions[r0].height = 26

    row = r0
    for p in POINTS:
        row += 1
        num, den = _series(tracked, p)
        cL = int(num.get(L, 0))
        cP = int(num.get(P, 0)) if P else 0
        cYA = int(num.get(YA, 0)) if YA else 0
        share = (cL / int(den.get(L, 0)) * 100) if den.get(L, 0) else 0
        mom = (cL - cP) / cP * 100 if (P and cP) else None
        yoy = (cL - cYA) / cYA * 100 if (YA and cYA) else None
        vals = [p["point"], p["airport"], ",".join(sorted(POINT_GATES[p["point"]])),
                (cYA if YA else "—"), (cP if P else "—"), cL,
                (f"{mom:+.0f}%" if mom is not None else "н/д"),
                (f"{yoy:+.0f}%" if yoy is not None else "н/д"),
                round(share, 1)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            c.border = BORDER
            c.font = BASE
            if i >= 2:
                c.alignment = CENTER
            if i == 1:
                c.font = BOLD
            if i == 3:
                c.font = Font(FONT, size=8)
            if i == 9:
                c.number_format = '0.0"%"'
            if i == 7 and mom is not None:
                c.font = GREEN if mom >= 0 else RED
            if i == 8 and yoy is not None:
                c.font = GREEN if yoy >= 0 else RED

    for i, w in enumerate([16, 5, 22, 11, 11, 11, 9, 9, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D5"


def build_reference(wb, whole):
    """Справочный лист: весь аэропорт (все гейты и рейсы, данные бота)."""
    ws = wb.create_sheet("Справочно — весь аэропорт")
    ws.cell(1, 1, "Справочно: весь аэропорт (все гейты и рейсы, данные бота)").font = TITLE_FONT
    ws.cell(2, 1, "Бот собирает все гейты аэропорта. Несопоставимо с листом точек "
                  "(там только гейты коллег). Доля точки = гейты точки ÷ все рейсы "
                  "терминала.").font = GREY
    months = sorted(set(whole["ym"]))[-6:]
    _hdr(ws, 4, ["Аэропорт"] + months)
    piv = (whole.pivot_table(index="airport", columns="ym", values="cgate",
                             aggfunc="count", fill_value=0)
           .reindex(columns=months, fill_value=0))
    r = 4
    for ap in [a for a in ["SVO", "VKO", "DME"] if a in piv.index]:
        r += 1
        ws.cell(r, 1, ap).font = BOLD
        ws.cell(r, 1).border = BORDER
        for j, m in enumerate(months):
            c = ws.cell(r, 2 + j, int(piv.loc[ap, m]))
            c.font = BASE
            c.border = BORDER
            c.alignment = CENTER
    r += 2
    ws.cell(r, 1, "Доля точки во ВСЁМ терминале (бот), %").font = BOLD
    r += 1
    _hdr(ws, r, ["Точка", "Аэр"] + months)
    r += 1
    for p in POINTS:
        num, den = _series(whole, p)
        ws.cell(r, 1, p["point"]).font = BOLD
        ws.cell(r, 2, p["airport"]).font = BASE
        ws.cell(r, 1).border = BORDER
        ws.cell(r, 2).border = BORDER
        ws.cell(r, 2).alignment = CENTER
        for j, m in enumerate(months):
            n = int(num.get(m, 0))
            d = int(den.get(m, 0))
            sh = (n / d * 100) if d else 0
            c = ws.cell(r, 3 + j, round(sh, 1) if n else None)
            c.number_format = '0.0"%"'
            c.font = BASE
            c.border = BORDER
            c.alignment = CENTER
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 6
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 9


def _hdr(ws, r, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def make_digest(tracked) -> list[str]:
    full = _full_months(tracked)
    if len(full) < 2:
        return ["Недостаточно полных месяцев для дайджеста."]
    last, prev = full[-1], full[-2]
    lines = [f"Дайджест за {last} (доля в наших гейтах терминала, базис коллег):"]
    ups, downs = [], []
    for p in POINTS:
        num, den = _series(tracked, p)
        cur = (num.get(last, 0) / den.get(last, 1) * 100) if den.get(last, 0) else 0
        pv = (num.get(prev, 0) / den.get(prev, 1) * 100) if den.get(prev, 0) else 0
        dm = cur - pv
        if abs(dm) >= 3:
            (ups if dm > 0 else downs).append(
                f"  {'▲' if dm > 0 else '▼'} {p['point']} ({p['airport']}): "
                f"{cur:.1f}%, {dm:+.1f} пп")
    if ups:
        lines.append("Рост:")
        lines += ups
    if downs:
        lines.append("Снижение:")
        lines += downs
    if not ups and not downs:
        lines.append("  Существенных изменений (≥3 пп) нет.")
    return lines


def build_report(out_path, dfrom=None, dto=None, _df=None):
    if _df is None:
        from src.analytics import load_all
        _df = load_all()
    if _df is None or _df.empty:
        raise SystemExit("Хранилище пусто — нет данных для отчёта.")
    d = _prep(_df, dfrom, dto)
    whole = _whole(d)
    tracked = _tracked(d)
    wb = Workbook()
    build_summary(wb, whole, tracked)
    build_points(wb, whole, tracked)
    build_reference(wb, whole)
    wb.save(out_path)
    return {
        "рейсов_всего": int(len(d)),
        "рейсов_бот": int(len(whole)),
        "рейсов_наши_гейты": int(len(tracked)),
        "период": f"{d['ym'].min()}..{d['ym'].max()}",
        "дайджест": make_digest(tracked),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="dfrom", default=None)
    ap.add_argument("--to", dest="dto", default=None)
    ap.add_argument("--out", default="Загрузка_гейтов_отчет.xlsx")
    args = ap.parse_args()
    info = build_report(args.out, args.dfrom, args.dto)
    print(f"Готово: {args.out}")
    for k, v in info.items():
        if k == "дайджест":
            print("\n" + "\n".join(v))
        else:
            print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
